"""
leitor.py
---------
Leitura das planilhas de controle de contratos, com diagnóstico.

Duas origens possíveis dentro do mesmo arquivo:

  1. Aba "Base de Dados" — leiaute nativo do painel (cabeçalho na linha 3,
     contratos a partir da linha 4, 65 colunas na ordem do XKEYS + 6 extras).
     É a fonte preferencial.

  2. Aba operacional do ano (ex.: "2027") — onde a secretaria efetivamente
     digita. A leitura aqui é DIRIGIDA POR CABEÇALHO: cada coluna é
     identificada pelo seu título, não pela posição. Isso torna a leitura
     imune à inserção de colunas (o leiaute 2027 inseriu "Modalidade de
     Licitação" na coluna 7, deslocando todas as seguintes).

Por que isso importa
--------------------
A aba "Base de Dados" é inteiramente montada por fórmulas que espelham a aba
operacional. Se o arquivo for salvo sem recálculo, essas fórmulas ficam sem
valor em cache e a aba é lida como VAZIA — a secretaria simplesmente some do
painel. Nesse caso o leitor cai para a aba operacional, onde os dados foram
digitados, e recompõe os campos calculáveis.
"""

from __future__ import annotations

import unicodedata
import warnings
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Aviso do openpyxl silenciado — explicação
# ---------------------------------------------------------------------------
# Ao ler uma planilha que usa listas suspensas cujo intervalo está em OUTRA
# aba (o caso das nossas planilhas, que puxam da aba "Listas"), o Excel grava
# essa validação num formato de EXTENSÃO que o openpyxl não interpreta. Ele
# então avisa: "Data Validation extension is not supported and will be
# removed".
#
# O "will be removed" refere-se à cópia que o openpyxl mantém em memória: se o
# arquivo fosse SALVO por ele, as listas suspensas se perderiam. Este sistema
# apenas LÊ planilhas — nunca grava — de modo que nada é alterado no arquivo da
# secretaria e nenhum dado é afetado.
#
# O aviso é suprimido apenas para manter o terminal legível.
warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported and will be removed",
    category=UserWarning,
    module="openpyxl",
)

# ---------------------------------------------------------------------------
# Ordem canônica das 65 chaves do painel (idêntica ao XKEYS do frontend).
# ---------------------------------------------------------------------------
XKEYS = [
    "num", "ficha", "processo", "contratada", "cnpj", "contrato_ano",
    "tipo_objeto", "objeto", "objeto_resumido", "regime", "codigo_acao",
    "tipo_empenho", "inicio_vig", "termino_vig", "unidade_gestora",
    "centro_custo", "programa", "natureza_despesa", "elemento_despesa",
    "fonte_recurso", "categoria_econ", "modalidade", "tipo_contratacao",
    "fiscal", "gestor", "valor_original", "data_assinatura", "situacao",
    "exercicio", "valor_mensal", "meses_vig", "valor_total_exerc",
    "dotacao_inicial", "dotacao_atualizada", "valor_empenhado",
    "valor_liquidado", "valor_pago", "saldo_empenhar", "saldo_liquidar",
    "saldo_pagar", "pct_empenhado", "pct_executado", "restos_pagar",
    "saldo_orc_unidade", "valor_suplementar", "valor_exec_mes",
    "valor_acumulado", "pct_fisico", "ultima_medicao", "proxima_medicao",
    "ultimo_pagamento", "proximo_pagamento", "qtde_aditivos",
    "valor_acrescido", "valor_suprimido", "novo_valor", "nova_vigencia",
    "ultima_justif", "dias_vencer", "vence_90", "vencido", "saldo_menor_20",
    "exec_maior_90", "sem_pag_60", "observacoes",
]

# Colunas 66–71 do leiaute 2027 (desdobramento do pago por fonte).
EXTRAS_2027 = {
    66: "pago_fonte_livre",
    67: "pago_fonte_vinculado",
    68: "pago_fonte_convenio",
    69: "pago_fonte_op_credito",
    70: "pago_fonte_emendas",
    71: "verificacao_pago_fonte",
}

CAMPOS_DATA = {
    "inicio_vig", "termino_vig", "data_assinatura", "ultima_medicao",
    "proxima_medicao", "ultimo_pagamento", "proximo_pagamento", "nova_vigencia",
}

CAMPOS_NUMERICOS = {
    "valor_original", "valor_mensal", "meses_vig", "valor_total_exerc",
    "dotacao_inicial", "dotacao_atualizada", "valor_empenhado",
    "valor_liquidado", "valor_pago", "saldo_empenhar", "saldo_liquidar",
    "saldo_pagar", "pct_empenhado", "pct_executado", "restos_pagar",
    "saldo_orc_unidade", "valor_suplementar", "valor_exec_mes",
    "valor_acumulado", "pct_fisico", "qtde_aditivos", "valor_acrescido",
    "valor_suprimido", "novo_valor", "dias_vencer", "meses_empenhados",
    "pago_fonte_livre", "pago_fonte_vinculado", "pago_fonte_convenio",
    "pago_fonte_op_credito", "pago_fonte_emendas",
}

# Leiaute nativo da aba "Base de Dados".
LINHA_CABECALHO_BASE = 3
PRIMEIRA_LINHA_BASE = 4
COL_CONTRATADA = 4
COL_OBJETO = 8

# ---------------------------------------------------------------------------
# Dicionário de cabeçalhos da aba operacional -> chave canônica.
# A comparação é feita por "contém", sobre o título normalizado (sem acento,
# minúsculo). A ordem importa: o primeiro padrão que casar vence.
# ---------------------------------------------------------------------------
PADROES_CABECALHO = [
    ("processo sei",                 "processo"),
    ("cpf/cnpj",                     "cnpj"),
    ("cnpj da contratada",           "cnpj"),
    ("contratada",                   "contratada"),
    ("tipo de objeto",               "tipo_objeto"),
    ("objeto do contrato",           "objeto"),
    ("modalidade",                   "modalidade"),
    ("regime de execucao",           "regime"),
    ("tipo de empenho",              "tipo_empenho"),
    ("inicio da vigencia",           "inicio_vig"),
    ("termino da vigencia",          "termino_vig"),
    ("valor mensal",                 "valor_mensal"),
    ("meses vigentes",               "meses_vig"),
    ("valor total no exercicio",     "valor_total_exerc"),
    ("meses empenhados",             "meses_empenhados"),
    ("valor empenhado",              "valor_empenhado"),
    ("saldo a empenhar",             "saldo_empenhar"),
    ("% empenhado",                  "pct_empenhado"),
    ("saldo orcamentario",           "saldo_orc_unidade"),
    ("valor a suplementar",          "valor_suplementar"),
    ("observacoes",                  "observacoes"),
    ("no do contrato/ano",           "contrato_ano"),
    ("unidade gestora",              "unidade_gestora"),
    ("fonte de recurso",             "fonte_recurso"),
    ("natureza da despesa",          "natureza_despesa"),
    ("elemento de despesa",          "elemento_despesa"),
    ("fiscal do contrato",           "fiscal"),
    ("gestor do contrato",           "gestor"),
    ("situacao",                     "situacao"),
    ("ficha",                        "ficha"),
]


def _norm(txt) -> str:
    if txt is None:
        return ""
    s = str(txt).strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def _texto(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _numero(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "").replace("%", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _data_iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    return s.split(" ")[0].split("T")[0]


def _converter(chave, bruto):
    if chave in CAMPOS_DATA:
        return _data_iso(bruto)
    if chave in CAMPOS_NUMERICOS:
        return _numero(bruto)
    return _texto(bruto)


# ---------------------------------------------------------------------------
# Aba "Base de Dados" (leiaute nativo, posicional)
# ---------------------------------------------------------------------------
def _achar_aba_base(wb):
    for nome in wb.sheetnames:
        n = _norm(nome)
        if n.startswith("base"):
            return nome
    return None


def _ler_base_de_dados(ws) -> tuple[list[dict], dict]:
    registros, rejeitadas = [], {"sem_contratada": 0, "sem_objeto": 0, "vazias": 0}
    varridas = 0
    n_cols = len(XKEYS) + len(EXTRAS_2027)

    for linha, valores in enumerate(
        ws.iter_rows(min_row=PRIMEIRA_LINHA_BASE, max_col=n_cols, values_only=True),
        start=PRIMEIRA_LINHA_BASE,
    ):
        contratada = _texto(valores[COL_CONTRATADA - 1]) if len(valores) >= COL_CONTRATADA else None
        objeto = _texto(valores[COL_OBJETO - 1]) if len(valores) >= COL_OBJETO else None
        if not contratada and not objeto:
            rejeitadas["vazias"] += 1
            continue
        varridas += 1
        if not contratada:
            rejeitadas["sem_contratada"] += 1
            continue
        if not objeto:
            rejeitadas["sem_objeto"] += 1
            continue

        reg = {}
        for idx, chave in enumerate(XKEYS):
            bruto = valores[idx] if idx < len(valores) else None
            reg[chave] = _converter(chave, bruto)
        for col, chave in EXTRAS_2027.items():
            bruto = valores[col - 1] if col - 1 < len(valores) else None
            reg[chave] = (_texto(bruto) if chave == "verificacao_pago_fonte"
                          else _numero(bruto))
        reg["_linha"] = linha
        registros.append(reg)

    return registros, {"linhas_com_conteudo": varridas, "rejeitadas": rejeitadas}


# ---------------------------------------------------------------------------
# Aba operacional do ano (leitura dirigida por cabeçalho)
# ---------------------------------------------------------------------------
def _achar_aba_operacional(wb):
    for nome in wb.sheetnames:
        n = nome.strip()
        if n.isdigit() and len(n) == 4:
            return nome
    ignorar = ("instruc", "lista", "base", "modelo", "ficha", "pretens")
    for nome in wb.sheetnames:
        if not any(_norm(nome).startswith(p) for p in ignorar):
            return nome
    return None


def _mapear_cabecalho(ws) -> tuple[int | None, dict[int, str]]:
    """Localiza a linha de cabeçalho e mapeia coluna -> chave canônica."""
    for linha, valores in enumerate(
        ws.iter_rows(min_row=1, max_row=20, max_col=60, values_only=True), start=1
    ):
        titulos = {c: _norm(v) for c, v in enumerate(valores, start=1)}
        texto = " | ".join(titulos.values())
        if "contratada" in texto and any(
            marca in texto for marca in
            ("objeto", "processo", "vigencia", "contrato/ano", "cpf/cnpj")
        ):
            mapa, usadas = {}, set()
            for col, t in titulos.items():
                if not t:
                    continue
                for padrao, chave in PADROES_CABECALHO:
                    if padrao in t and chave not in usadas:
                        mapa[col] = chave
                        usadas.add(chave)
                        break
            return linha, mapa
    return None, {}


def _ler_operacional(ws) -> tuple[list[dict], dict]:
    linha_cab, mapa = _mapear_cabecalho(ws)
    if not linha_cab:
        return [], {"erro": "cabeçalho não localizado na aba operacional"}

    col_contratada = next((c for c, k in mapa.items() if k == "contratada"), None)
    col_objeto = next((c for c, k in mapa.items() if k == "objeto"), None)
    if not col_contratada:
        return [], {"erro": "coluna 'Contratada' não localizada"}

    max_col = max(mapa) if mapa else 30
    registros, rejeitadas = [], {"sem_contratada": 0, "sem_objeto": 0, "vazias": 0}
    varridas = 0

    for linha, valores in enumerate(
        ws.iter_rows(min_row=linha_cab + 1, max_col=max_col, values_only=True),
        start=linha_cab + 1,
    ):
        pega = lambda c: valores[c - 1] if c and c - 1 < len(valores) else None
        contratada = _texto(pega(col_contratada))
        objeto = _texto(pega(col_objeto))
        if not contratada and not objeto:
            rejeitadas["vazias"] += 1
            continue
        varridas += 1
        if not contratada:
            rejeitadas["sem_contratada"] += 1
            continue

        reg = {k: None for k in XKEYS}
        for col, chave in mapa.items():
            reg[chave] = _converter(chave, pega(col))

        if not reg.get("objeto"):
            if reg.get("tipo_objeto"):
                reg["objeto"] = reg["tipo_objeto"]
            else:
                rejeitadas["sem_objeto"] += 1
                continue

        _recompor_calculados(reg)
        reg["_linha"] = linha
        registros.append(reg)

    return registros, {
        "linha_cabecalho": linha_cab,
        "colunas_reconhecidas": len(mapa),
        "linhas_com_conteudo": varridas,
        "rejeitadas": rejeitadas,
    }


def _recompor_calculados(reg: dict) -> None:
    """
    Recompõe os campos que a planilha calcula por fórmula, para o caso de o
    arquivo ter sido salvo sem recálculo (fórmulas sem valor em cache).
    """
    mensal = reg.get("valor_mensal")
    meses = reg.get("meses_vig")
    total = reg.get("valor_total_exerc")
    empenhado = reg.get("valor_empenhado")

    if total is None and mensal is not None and meses is not None:
        total = round(mensal * meses, 2)
        reg["valor_total_exerc"] = total

    if reg.get("saldo_empenhar") is None and total is not None:
        reg["saldo_empenhar"] = round(total - (empenhado or 0.0), 2)

    if reg.get("pct_empenhado") is None and total:
        reg["pct_empenhado"] = (empenhado or 0.0) / total

    reg["valor_liquidado"] = reg.get("valor_liquidado") or 0.0
    reg["valor_pago"] = reg.get("valor_pago") or 0.0
    reg["saldo_liquidar"] = empenhado
    reg["saldo_pagar"] = reg.get("saldo_pagar") or 0.0
    reg["pct_executado"] = reg.get("pct_executado") or 0.0


# ---------------------------------------------------------------------------
# Detecção de planilha não recalculada
# ---------------------------------------------------------------------------
def _formulas_sem_cache(caminho: Path, aba: str) -> int:
    """
    Conta, nas colunas que identificam o contrato, células com fórmula e sem
    valor em cache — assinatura de um arquivo salvo sem recálculo.
    A varredura é restrita para manter o processamento rápido.
    """
    try:
        wf = openpyxl.load_workbook(caminho, data_only=False, read_only=True)
        wv = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    except Exception:  # noqa: BLE001
        return 0
    try:
        if aba not in wf.sheetnames:
            return 0
        sf, sv = wf[aba], wv[aba]
        pendentes = 0
        linhas_f = sf.iter_rows(min_row=PRIMEIRA_LINHA_BASE, max_row=PRIMEIRA_LINHA_BASE + 60,
                                min_col=1, max_col=COL_OBJETO, values_only=True)
        linhas_v = sv.iter_rows(min_row=PRIMEIRA_LINHA_BASE, max_row=PRIMEIRA_LINHA_BASE + 60,
                                min_col=1, max_col=COL_OBJETO, values_only=True)
        for lf, lv in zip(linhas_f, linhas_v):
            for f, v in zip(lf, lv):
                if isinstance(f, str) and f.startswith("=") and v is None:
                    pendentes += 1
        return pendentes
    finally:
        wf.close()
        wv.close()


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------
def ler_planilha(caminho: Path) -> tuple[list[dict], dict]:
    """
    Lê uma planilha de controle de contratos.

    Ordem de tentativa:
      1. Aba "Base de Dados" (leiaute nativo, posicional).
      2. Aba operacional do ano, por cabeçalho.
      3. Qualquer outra aba, por cabeçalho — vence a que trouxer mais contratos.

    O workbook é SEMPRE fechado ao final. Em modo somente-leitura o openpyxl
    mantém o arquivo aberto, e no Windows um arquivo aberto não pode ser
    excluído — o que impedia a substituição do pacote sem reiniciar o servidor.
    """
    caminho = Path(caminho)
    diag: dict = {"abas": [], "origem": None, "contratos": 0,
                  "motivo": None, "detalhe": {}, "tentativas": []}

    nome_base = None
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    try:
        diag["abas"] = list(wb.sheetnames)

        # 1) Aba "Base de Dados"
        nome_base = _achar_aba_base(wb)
        if nome_base:
            regs, det = _ler_base_de_dados(wb[nome_base])
            diag["tentativas"].append({"aba": nome_base, "modo": "nativo",
                                       "contratos": len(regs)})
            if regs:
                diag.update(origem=f"aba '{nome_base}'", contratos=len(regs),
                            detalhe=det)
                return regs, diag
            diag["detalhe"]["base_de_dados"] = det

        # 2) e 3) Demais abas, por cabeçalho — a melhor vence.
        melhor: tuple[list[dict], dict, str] | None = None
        preferida = _achar_aba_operacional(wb)
        ordem = ([preferida] if preferida else []) + [
            n for n in wb.sheetnames if n != preferida and n != nome_base
        ]

        for nome in ordem:
            if nome is None:
                continue
            try:
                regs, det = _ler_operacional(wb[nome])
            except Exception:  # noqa: BLE001
                continue
            diag["tentativas"].append({"aba": nome, "modo": "cabeçalho",
                                       "contratos": len(regs)})
            if regs and (melhor is None or len(regs) > len(melhor[0])):
                melhor = (regs, det, nome)
            if melhor and nome == preferida:
                break   # a aba do ano tem prioridade quando traz dados

        if melhor:
            regs, det, nome = melhor
            diag.update(origem=f"aba '{nome}' (por cabeçalho)",
                        contratos=len(regs))
            diag["detalhe"]["operacional"] = det
            return regs, diag
    finally:
        wb.close()

    # Nada encontrado: explica o porquê, nomeando o que foi testado.
    testadas = ", ".join(
        f"'{t['aba']}' ({t['contratos']})" for t in diag["tentativas"]
    ) or "nenhuma"

    if nome_base:
        pendentes = _formulas_sem_cache(caminho, nome_base)
        diag["detalhe"].setdefault("base_de_dados", {})["formulas_sem_valor"] = pendentes
        if pendentes:
            diag["motivo"] = (
                f"A aba '{nome_base}' é montada por fórmulas e o arquivo foi "
                f"salvo sem recálculo ({pendentes} fórmulas sem valor). "
                f"Abra a planilha no Excel, salve e envie novamente. "
                f"Abas testadas: {testadas}."
            )
        else:
            diag["motivo"] = (
                f"Nenhuma aba trouxe contratos preenchidos. "
                f"Abas testadas (contratos lidos): {testadas}."
            )
    elif diag["tentativas"]:
        diag["motivo"] = (
            f"Nenhuma aba trouxe contratos preenchidos. "
            f"Abas testadas (contratos lidos): {testadas}."
        )
    else:
        diag["motivo"] = (
            f"Nenhuma aba foi reconhecida como base de contratos. Abas do "
            f"arquivo: {', '.join(diag['abas']) or '(nenhuma)'}. Verifique se "
            f"existe a aba 'Base de Dados' ou a aba do ano com as colunas "
            f"'Contratada' e 'Objeto do Contrato'."
        )
    return [], diag
